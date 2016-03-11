import pandas as pd
import numpy as np
import os
from pprint import pprint
from xlwings import Workbook, Range, Sheet
from openpyxl import load_workbook

from iterate_in_array import fill_array_with_excel_formulas
from iterate_in_array import fill_array_with_excel_formulas_based_on_is_forecast   

from import_specification import get_all_input_variables, get_array_and_support_variables

                         
###########################################################################
## Export to Excel workbook (using xlwings/pywin32 or openpyxl)
###########################################################################

def save_xl_using_xlwings(file):  
    wb = Workbook(file)
    wb.save()

def write_array_to_xl_using_xlwings(ar, file, sheet):
    # Note: if file is opened In Excel, it must be first saved before writing 
    #       new output to it, but it may be left open in Excel application. 
    wb = Workbook(file)
    Sheet(sheet).activate()

    def nan_to_empty_str(x):
        return '' if type(x) == float and np.isnan(x) else x

    Range(sheet, 'A1').value = [[nan_to_empty_str(x) for x in row] for row in ar]
    wb.save()

#--------------------------------------------------------------------------
# Not tested

def change_extension(file):
    """
    >>> change_extension("spec.xls")
    'spec.xlsx'
    """
    return os.path.splitext(file) + ".xlsx"

def iterate_over_array(ar):
    for i, row in enumerate(ar):       
         for j, val in enumerate(row):
                yield i, j, val
                
def write_array_to_xlsx_using_openpyxl(ar, file, sheet):  
    wb = load_workbook(file)
    ws = wb.get_sheet_by_name(sheet)
    for i, j, val in iterate_over_array(ar):
        ws.cell(row = i, column = j).value = val
    new_filename = change_extension(file)
    wb.save(new_filename) 
    
#--------------------------------------------------------------------------
    
###########################################################################
## Dataframe manipulation
###########################################################################

def make_empty_df(index_, columns_):
    df = pd.DataFrame(index=index_, columns=columns_)
    return  df 

def subset_df(df, var_list):
    try:
        return df[var_list]
    except KeyError:        
        pprint ([x for x in var_list if x not in df.columns.values])
        raise KeyError("*var_list* contains variables outside *df* column names." +  
                       "\nCannot perform subsetting like df[var_list]")
    except:
        print ("Error handling dataframe:", df)
        raise ValueError
      
def make_df_before_equations(data_df, controls_df, equations_dict, var_group):
    """
    Return a dataframe containing data, controls and a placeholder for new 
    varaibales derived in equations.
    """    
    IS_FORECAST_LABEL = 'is_forecast'
    
    # assign 'is_forecast' to dataframes
    data_df[IS_FORECAST_LABEL] = 0 
    controls_df[IS_FORECAST_LABEL] = 1
     
    # concat data and control *df*
    df = data_df.combine_first(controls_df)
    
    
    # *df2* is a placeholder for equation-derived variables 
    df3 = make_empty_df(data_df.index, var_group['eq'])
    # add *df3* to *df*. 
    df = pd.merge(df, df3, left_index = True, right_index = True, how = 'left')
    
    # reorganise rows
    
    var_list = var_group['data'] +  var_group['eq'] + var_group['control'] + [IS_FORECAST_LABEL]
    return subset_df(df, var_list)


###########################################################################
## Array manipulations
###########################################################################

from iterate_in_array import get_variable_rows_as_dict

def make_array_before_equations(df):
    """
    Convert dataframe to array, decorate with extra top row an extra left-side columns.
    Returns array and pivot column number. Pivot column contains variable labels.
    """
    ar = df.as_matrix().transpose().astype(object)
    
    # add variable labels as a first column in *ar*
    labels = df.columns.tolist()
    ar = np.insert(ar, 0, labels, axis = 1)
    pivot_col = 0

    # add years as first row in *ar*
    years = [""] + df.index.astype(str).tolist()
    ar = np.insert(ar, 0, years, axis = 0)
    
    return ar, pivot_col

###### After equations

def insert_empty_row_before_variable(ar, var_name, pivot_col, start_cell_text = ""):
    variables_dict = get_variable_rows_as_dict(ar, pivot_col)
    row_position = variables_dict[var_name] 
    ar = np.insert(ar, row_position, "", axis = 0) 
    ar[row_position, 0] = start_cell_text 
    return ar

def insert_column(ar, pivot_col, datagen_func):
    column_values = [datagen_func(x) for x in ar[:, pivot_col]]
    ar = np.insert(ar, 0, column_values, axis = 1)
    return ar, pivot_col + 1   

def append_row_to_array(ar):
    row = [["" for x in ar[0,:]]]    
    return np.append(ar, row, axis = 0)
    
def add_equations_to_array (ar, pivot_col, eq_list):    
    for eq in eq_list:
        ar = append_row_to_array(ar)
        ar[-1, pivot_col] = eq
    return ar

    
###########################################################################
## Main entry point
###########################################################################

                         
def get_resulting_workbook_array_for_make(abs_filepath, slim = True):

    # Get model specification
    data_df, controls_df, equations_dict, var_group, var_desc_dict, eq_list = get_all_input_variables(abs_filepath) 
     
    # Get array before formulas
    df = make_df_before_equations(data_df, controls_df, equations_dict, var_group)
    ar, pivot_col = make_array_before_equations(df) 
    
    if not slim:
        # Decorate with extra columns ---------------------------------------------
        def null(x):    
           return ""
           
        def get_var_desc(varname):
           if varname in var_desc_dict.keys():
               return var_desc_dict[varname]
           else:
               return ""       
           
        ar, pivot_col = insert_column(ar, pivot_col, get_var_desc)
        ar, pivot_col = insert_column(ar, pivot_col, null)              
       
        # Decorate with extra empty rows 
        def insert_row(t, gen):
            # t is (varname, start_cell_text)
            return insert_empty_row_before_variable(ar, t[0], 
                                                    pivot_col, next(gen) + t[1])        
        def yield_chapter_numbers():
            for i in [1,2,3,4]:
                 yield str(i)  
                                                   
        gen = yield_chapter_numbers()
        
        dec_dict = { "data": (var_group['data'][0],    ". ИСХОДНЫЕ ДАННЫЕ И ПРОГНОЗ"),
                     "ctrl": (var_group['control'][0], ". УПРАВЛЯЮЩИЕ ПАРАМЕТРЫ")}                      
        if var_group['eq']:
             dec_dict['eq'] = (var_group['eq'][0],      ". ПЕРЕМЕННЫЕ ИЗ УРАВНЕНИЙ")
                        
        ar = insert_row(dec_dict['data'], gen)
        if var_group['eq']:        
            ar = insert_row(dec_dict['eq'], gen)
        ar = insert_row(dec_dict['ctrl'], gen)
                        
        # -------------------------------------------------------------------------    
       
    ar = fill_array_with_excel_formulas(ar, equations_dict, pivot_col)
    
    if not slim:
        ar = append_row_to_array(ar)
        ar[-1,0] = next(gen) + ". УРАВНЕНИЯ"
        ar = add_equations_to_array (ar, pivot_col, eq_list)
        
    return ar

# pivot_col = 2 is standard output of --make --fancy
def update_xl_model(abs_filepath, sheet, pivot_col = 2): 
    save_xl_using_xlwings(abs_filepath) 
    ar, equations_dict = get_array_and_support_variables(abs_filepath, sheet, pivot_col)         
    ar = fill_array_with_excel_formulas_based_on_is_forecast(ar, equations_dict, pivot_col)    
    print("\nResulting Excel sheet as array:")     
    print(ar)  
    write_array_to_xl_using_xlwings(ar, abs_filepath, sheet)
    
def make_xl_model(abs_filepath, sheet, slim, use_dataset): 
    ar = get_resulting_workbook_array_for_make(abs_filepath, slim)    
    print("\nResulting Excel sheet as array:")     
    print(ar)
    write_array_to_xl_using_xlwings(ar, abs_filepath, sheet)
    
if __name__ == '__main__':
    
#    for fn in ['spec.xls', 'spec2.xls']:
#       abs_filepath = os.path.abspath(fn)
#       sheet = 'model'
#       make_xl_model(abs_filepath, sheet, slim = False)
    
    abs_filepath = os.path.abspath('spec.xls') 
    sheet = 'model'
    make_xl_model(abs_filepath, sheet, slim = False)     
    #update_xl_model(abs_filepath, sheet, pivot_col = 2)
