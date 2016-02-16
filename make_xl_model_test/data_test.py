# -*- coding: utf-8 -*-
"""
Test for procedure to create resulting Excel sheet based on data, formulas and control parameters. 

Key elements:
  отчетные данные (data) - ряды данных с названиями переменных 
  формулы (formulas) - символьные выражения, которые определяют прогнозные значения следующего периода
  параметры (controls) - управляющие параметры, которые используются при расчете прогнозных значений c помощью формул, может задаваться как p_*
  размещение (layout) - расположение рядов данных в итоговом Excel файле  
  
Limitations:
  1. no new variables created in formulas
  2. not writing variable text descriptions, only variable labels
"""

import pandas as pd
dff = pd.DataFrame
tsf = pd.TimeSeries

# --------------------------------------------------------------------
# Data
obs_years = [2010, 2011,  2012,  2013, 2014]
obs_x     = [100,   101, 102.5, 100.8, 95.5]
obs_y     = [3360, 2700,   500,  1200, 4800]
data = pd.DataFrame({'x': obs_x, 
                     'y': obs_y}, index = obs_years)
print("\nData:", data)

# --------------------------------------------------------------------
# Controls
forecast_years = [2015, 2016] #[obs_years[-1] + x for x in range(1,3)] 
y_rog = [1.1,  1.2]
x_fut = [102.5, 105]
controls = pd.DataFrame({'y_rog': y_rog,
                         'x_fut': x_fut}, index = forecast_years)
print("\nControls:", controls)

# --------------------------------------------------------------------
# Formulas
eq1 = "x[t] = x_fut[t]"
eq2 = "y[t] = y[t-1] * y_rog[t]"
# LIMITATION: variable z[t] must be in data before use, cannot add new variables by equationas of now
# eq3 = "z[t] = x[t] / y [t]" 
formulas = [eq1, eq2]
print("\nFormulas:", formulas)

# --------------------------------------------------------------------
# Resulting dataframe with values
all_years = obs_years + forecast_years
is_forecast = [0 for x in obs_years] + [1 for x in forecast_years]
data = data.reindex(index = all_years)
controls = controls.reindex(index = all_years)


iterator = zip([(t, isf) for t, isf in enumerate(is_forecast)], all_years)
for t, year in enumerate(all_years):
    if is_forecast[t]:
        # mimic formulas:
        data.x[year] = controls.x_fut[year]
        data.y[year] = data.y[year-1] * controls.y_rog[year]

        
output_layout = {'sheet': 'result',
                 'upper_left_corner': 'B2',
                 'variable_list': ['x', 'y', 'x_fut', 'y_rog']
                }

var_list = output_layout['variable_list'] 
out = data.join(controls)[var_list]
print("\nResulting dataframe with values:", out)




def write_sheet(df, sheet_name, writer):
    df.transpose().to_excel(writer, sheet_name)

def write_formulas(formulas_as_list, writer):
    pd.DataFrame(formulas_as_list).to_excel(writer, 'formulas', 
                 index = False)
def read_df(file, sheet_name):
    return pd.read_excel(file, sheet_name).transpose()
def read_formulas(file):
    df = pd.read_excel(file, 'formulas', index_col=None)
    return df[0].values.tolist()

# write to file 
writer = pd.ExcelWriter('testfile.xls')
write_sheet(data, 'data', writer)
write_sheet(controls, 'controls', writer)
write_formulas(formulas, writer)
write_sheet(out, 'result', writer)
writer.save()

writer = pd.ExcelWriter('testfile.xls')



# read from file 
REF_FILE = "ref_file.xls"
assert read_df(REF_FILE, 'data').equals(data)
assert read_df(REF_FILE, 'controls').equals(controls)
assert read_df(REF_FILE, 'result').equals(out)
assert read_formulas(REF_FILE) == formulas


#import xlrd
#book = xlrd.open_workbook(REF_FILE)
#sh = book.sheet_by_name("result")
#print ("Row 1:", sh.row(1))
#for rx in range(sh.nrows):
#    print (sh.row(rx))
    # Refer to docs for more details.
    # Feedback on API is welcomed.





from openpyxl import load_workbook
wb = load_workbook(filename='ref_file.xlsx')
ws = wb['result'] 

# Start from the first cell. Rows and columns are zero indexed.
rn = 0
for row in ws.rows:
    cn = 0 
    for cell in row:
        print(rn, cn, cell.value)
        cn += 1
    rn +=1

# Start from the first cell. Rows and columns are zero indexed.
def _iter_rows(file, sheet): 
    for row in load_workbook(filename=file)[sheet].rows:
        yield ([cell.value for cell in row])


z = pd.DataFrame([x for x in _iter_rows('ref_file.xlsx', 'result')])
df = z.iloc[1:,1:]
df.index = z.iloc[1:,0]
df.columns = z.iloc[0,1:]
df.to_excel("df.xlsx")



"""
write fixtures to xls file on different sheets: data, controls, formulas
write 'out' to same file in 'upper_left_corner' location
change formulas, save as reference file - ref.xlsx
read formulas 
"""

