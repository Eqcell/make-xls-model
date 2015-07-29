import pandas as pd
import numpy as np
import os
from pprint import pprint
from xlwings import Workbook, Range, Sheet
from openpyxl import load_workbook

from formula_parser import make_eq_dict
from iterate_in_array import fill_array_with_excel_formulas   

   
###########################################################################
## Generic import from Excel workbook (using pandas)
###########################################################################

def read_sheet(filename_, sheet_, header_):    
    return pd.read_excel(filename_, sheetname=sheet_, header = header_).transpose()
    
def read_df(filename_, sheet_):    
    return read_sheet(filename_, sheet_, 0)
    
def read_col(filename_, sheet_):    
    return read_sheet(filename_, sheet_, None).values.tolist()[0]  
    
###########################################################################
## Import model specification, make it available as dict or tuple 
###########################################################################
    
def get_data_df(file):
    return read_df(file, 'data') 

def get_controls_df(file):
    return read_df(file, 'controls') 

def get_equations_dict(file):
    list_of_strings = read_col(file, 'equations')
    # todo: 
    #     parse_to_formula_dict must:
    #        - control left side of equations
    return make_eq_dict(list_of_strings)
    
def get_spec_as_dict(file):   
    return   { 'data': get_data_df(file)    
       ,   'controls': get_controls_df(file) 
       ,  'equations': get_equations_dict(file)
       }
       
def get_spec_as_tuple(file): 
    s = get_spec_as_dict(file)
    return s['data'], s['controls'], s['equations']
 
###########################################################################
## Input validation
###########################################################################

def get_input_variables(abs_filepath):
    data_df, controls_df, equations_dict = get_spec_as_tuple(abs_filepath) 
    var_group = get_variable_names_by_group(data_df, controls_df, equations_dict)
    return data_df, controls_df, equations_dict, var_group

def list_array(a):
    return  " ".join(str(x) for x in a)

def validate_input_from_sheets(abs_filepath):
    # Get model specification 
    data_df, controls_df, equations_dict, var_group = get_input_variables(abs_filepath)  
    
    validate_continious_year(data_df, controls_df)
    validate_coverage_by_equations(var_group, equations_dict)
    
def validate_continious_year(data_df, controls_df):
    # Data and controls must have continious timeline
    years1 = data_df.index.tolist() 
    years2 = controls_df.index.tolist()
    timeline = years1 + years2
    ref_timeline = [x for x in range(min(timeline), max(timeline) + 1)]
    if not timeline == ref_timeline:
        raise ValueError("Timeline derived from 'data' and 'controls' is not continious." +
            "\nData timeline: "      + list_array(years1)   +
            "\nControls timeline: "  + list_array(years2)   +
            "\nResulting timeline: " + list_array(timeline) +
            "\nExpected timeline: "  + list_array(ref_timeline)
            )

def validate_coverage_by_equations(var_group, equations_dict):    
    # Validate coverage of data_df with equations
    data_orphan_vars = [v for v in var_group["data"] if v not in equations_dict.keys()]
    if data_orphan_vars:
        print(data_orphan_vars)
        raise ValueError("All data variables must be covered by equations." +
                         "\nNot covered: " + list_array(data_orphan_vars))
                         
###########################################################################
## Export to Excel workbook (using xlwings/pywin32 or openpyxl)
###########################################################################

def write_array_to_xl_using_xlwings(ar, file, sheet):  
    # Note: if file is opened In Excel, it must be first saved before writing 
    #       new output to it, but it may be left open in Excel application. 
    wb = Workbook(file)
    Sheet(sheet).activate()        
    Range(sheet, 'A1').value = ar.astype(str)    
    wb.save()

#--------------------------------------------------------------------------
# Not tested

def change_extension(file):
    return os.path.splitext(file) + ".xlsx"

def iterate_over_array(ar):
    for i, row in enumerate(ar):       
         for j, val in enumerate(row):
                yield i, j, to_float(val)
                
def write_array_to_xlsx_using_openpyxl(ar, file, sheet):  
    wb = load_workbook(file)
    ws = wb.get_sheet_by_name(sheet)
    for i, j, val in iterate_over_array(ar):
        ws.cell(row = i, column = j).value = val
    new_filename = change_extension(file)
    wb.save(new_filename) 
    
#--------------------------------------------------------------------------
    
###########################################################################
## Grouped variables
###########################################################################
   
#def print_variable_names_by_group(group_dict):    
#    print ("Data vars:", group_dict['data'])
#    print ("Control vars:", group_dict['control'])
#    print ("Equation-derived vars:", group_dict['eq'])    
    
def get_variable_names_by_group(data_df, controls_df, equations_dict):
    """
    Obtain non-overlapping variable labels grouped into data, control 
    and equation-derived variables.    
    """
    
    # all variables from controls_df must persist in var_list (group 1 of variables)
    g1 = controls_df.columns.values.tolist()
    
    # group 2: variables in data_df not listed in control_df
    dvars = data_df.columns.values.tolist()
    g2 = [d for d in dvars if d not in g1]

    # group 3: variables on leftside of equations not listed in group 1 and 2
    evars = equations_dict.keys()
    g3 = [e for e in evars if e not in g1 + g2]
    
    return {'control': g1, 'data': g2, 'eq': g3}

###########################################################################
## Dataframe manipulation
###########################################################################

def make_empty_df(index_, columns_):
    df = pd.DataFrame(index=index_, columns=columns_)
    return  df 

def subset_df(df, var_list):
    try:
        return df[var_list]
    except KeyError:        
        pprint ([x for x in var_list if x not in df.columns.values])
        raise KeyError("*var_list* contains variables outside *df* column names." +  
                       "\nCannot perform subsetting like df[var_list]")
    except:
        print ("Error handling dataframe:", df)
        raise ValueError
      
def make_df_before_equations(data_df, controls_df, equations_dict):
    """
    Return a dataframe containing data, controls and a placeholder for new 
    varaibales derived in equations.
    """    
    
    var_group = get_variable_names_by_group(data_df, controls_df, equations_dict)
    
    # concat data and control *df*
    df1 = data_df
    df2 = controls_df
    df = pd.concat([df1, df2])    
    
    # *df2* is a placeholder for equation-derived variables 
    df3 = make_empty_df(data_df.index, var_group['eq'])
    # add *df3* to *df*. 
    df = pd.merge(df, df3, left_index = True, right_index = True, how = 'left')
    
    # reorganise rows
    var_list = var_group['data'] +  var_group['eq'] + var_group['control']
    return subset_df(df, var_list)


###########################################################################
## Array manipulations
###########################################################################

from iterate_in_array import get_variable_rows_as_dict

def insert_empty_row_before_variable(ar, var_name, pivot_col, top_value = ""):
    variables_dict = get_variable_rows_as_dict(ar, pivot_col)
    row_position = variables_dict[var_name] 
    ar = np.insert(ar, row_position, "", axis = 0) 
    ar[row_position, 0] = top_value     
    return ar

def proxy_dg(x):
    # return "NAME OF " + str(x)
    return "" 

def insert_column(ar, pivot_col, datagen_func):
    column_values = [datagen_func(x) for x in ar[:, pivot_col]]
    ar = np.insert(ar, 0, column_values, axis = 1)
    return ar, pivot_col + 1     

def make_array_before_equations(df):
    """
    Convert dataframe to array, decorate with extra top row an extra left-side columns.
    Returns array and pivot column number. Pivot column contains variable labels.
    """
    ar = df.as_matrix().transpose().astype(object)
    
    # add variable labels as a first column in *ar*
    labels = df.columns.tolist()
    ar = np.insert(ar, 0, labels, axis = 1)
    pivot_col = 0

    # add years as first row in *ar*
    years = [""] + df.index.astype(str).tolist()
    ar = np.insert(ar, 0, years, axis = 0)
    
    return ar, pivot_col
    

###########################################################################
## Main entry point
###########################################################################
                         
def get_resulting_workbook_array(abs_filepath):

    # Require all data is covered by equations
    validate_input_from_sheets(abs_filepath)
    
    # Get model specification 
    data_df, controls_df, equations_dict, var_group = get_input_variables(abs_filepath)    
    
    # Get array before formulas
    df = make_df_before_equations(data_df, controls_df, equations_dict)
    ar, pivot_col = make_array_before_equations(df) 
    
    # Decorate with extra column
    ar, pivot_col = insert_column(ar, pivot_col, proxy_dg)    
    
    # Decorate with extra empty rows
    def insert_row(var_name, top_value):
        return insert_empty_row_before_variable(ar, var_name, pivot_col, top_value)

    ar = insert_row(var_group['data'][0],    "Исходные данные и прогноз")
    if var_group['eq']:
        ar = insert_row(var_group['eq'][0],  "Переменные из уравнений")
    ar = insert_row(var_group['control'][0], "Управляющие параметры")
       
    # Fill array with formulas
    # Todo: fillable_var_list is effectively everything that appears on the left side of equations
    #       must compare *equations_dict* and *fillable_var_list* 
       
    fillable_var_list = var_group['data'] + var_group['eq']
    ar = fill_array_with_excel_formulas(ar, equations_dict, fillable_var_list, pivot_col)
    return ar

def update_xl_model(abs_filepath, sheet): 
    pass
    
def make_xl_model(abs_filepath, sheet): 
    ar = get_resulting_workbook_array(abs_filepath)    
    print("\nResulting Excel sheet as array:")     
    print(ar) 
    write_array_to_xl_using_xlwings(ar, abs_filepath, sheet)
    
if __name__ == '__main__':
    
    import os
    for fn in ['spec.xls', 'spec2.xls']:
       abs_filepath = os.path.abspath(fn)
       sheet = 'model'
       make_xl_model(abs_filepath, sheet)